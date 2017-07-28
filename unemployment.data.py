# -*- coding: utf-8 -*-
"""
Created on Thu Jun 15 15:30:42 2017

@author: finnjj
"""
import openpyxl
import pandas as pd
import numpy as np
from pandas import DataFrame, Series

datacodeDict = { 
    'AL' : 'Alabama',
    'AK' : 'Alaska',
    'AZ' : 'Arizona',
    'AR' : 'Arkansas',
    'CA' : 'California',
    'CO' : 'Colorado',
    'CT' : 'Connecticut',
    'DE' : 'Delaware',
    'FL' : 'Florida',
    'GA' : 'Georgia',
    'HI' : 'Hawaii',
    'ID' : 'Idaho',
    'IL' : 'Illinois',
    'IN' : 'Indiana',
    'IA' : 'Iowa',
    'KS' : 'Kansas',
    'KY' : 'Kentucky',
    'LA' : 'Louisiana',
    'ME' : 'Maine',
    'MD' : 'Maryland',
    'MA' : 'Massachusetts',
    'MI' : 'Michigan',
    'MN' : 'Minnesota',
    'MS' : 'Mississippi',
    'MO' : 'Missouri',
    'MT' : 'Montana',
    'NE' : 'Nebraska',
    'NV' : 'Nevada',
    'NH' : 'New Hampshire',
    'NJ' : 'New Jersey',
    'NM' : 'New Mexico',
    'NY' : 'New York',
    'NC' : 'North Carolina',
    'ND' : 'North Dakota',
    'OH' : 'Ohio',
    'OK' : 'Oklahoma',
    'OR' : 'Oregon',
    'PA' : 'Pennsylvania',
    'RI' : 'Rhode Island',
    'SC' : 'South Carolina',
    'SD' : 'South Dakota',
    'TN' : 'Tennessee',
    'TX' : 'Texas',
    'UT' : 'Utah',
    'VT' : 'Vermont',
    'VA' : 'Virginia',
    'WA' : 'Washington',
    'WV' : 'West Virginia',
    'WI' : 'Wisconsin',
    'WY' : 'Wyoming'}

year = []
statename = []
statecode = []
countyname = []
countycode = []
totallabor = []
unemployed = []
unemp_rate = []

year_st = []
statename_st = []
statecode_st = []
totallabor_st = []
unemployed_st = []
unemp_rate_st = []

def appendlists(yr, tot, unemp, rate):
    if (isinstance(ws.cell(row=i, column=tot).value, (int, float))) and (isinstance(ws.cell(row=i, column=unemp).value, (int, float))) and (isinstance(ws.cell(row=i, column=rate).value, (int, float))):
        #print(st_code + " " + co_code)                          # debug
        #print(":", ws.cell(row=i, column=rate).value, ":")      # debug
        #print(type(ws.cell(row=i, column=tot).value))           # debug
        #print(type(ws.cell(row=i, column=unemp).value))         # debug
        if (co_code != ''):
            year.append(yr)
            statename.append(st_name)
            statecode.append(st_code)
            countyname.append(co_name)
            countycode.append(co_code)
            totallabor.append(ws.cell(row=i, column=tot).value)
            unemployed.append(ws.cell(row=i, column=unemp).value)
            unemp_rate.append(round(float(ws.cell(row=i, column=rate).value) / 100, 3))
        else:
            year_st.append(yr)
            statename_st.append(st_name)
            statecode_st.append(st_code)
            totallabor_st.append(ws.cell(row=i, column=tot).value)
            unemployed_st.append(ws.cell(row=i, column=unemp).value)
            unemp_rate_st.append(round(float(ws.cell(row=i, column=rate).value) / 100, 3))
            


wb = openpyxl.load_workbook('Unemployment.counties.2007-15.xlsx')
ws = wb.get_sheet_by_name('Unemployment Med HH Inc ')
for i in range(9,3204):
    if ((ws.cell(row=i, column=2).value) == 'PR'):
        break
    if (not ws.cell(row=i, column=1).value) or ((ws.cell(row=i, column=2).value) == 'DC'):
        continue
    #print(type(ws.cell(row=i, column=1).value))                     # debug
    #if ((ws.cell(row=i, column=1).value % 1000) == 0):              # doesn't work
    if ((ws.cell(row=i, column=1).value[2:]) == '000'):
        st_name = datacodeDict.get(ws.cell(row=i, column=2).value)
        #st_code = str(ws.cell(row=i, column=1).value / 1000).zfill(2)
        st_code = ws.cell(row=i, column=1).value[0:2]
        co_code = ''
        co_name = ''
    else:
        co_code = str(ws.cell(row=i, column=1).value).zfill(5)
        #removestr = ' County, ' +  ws.cell(row=i, column=2).value        # doesn't work
        #co_name = str(ws.cell(row=i, column=3).value).rstrip(removestr)  # doesn't work 
        co_name = str(ws.cell(row=i, column=3).value)
        where = co_name.find('County,')
        if (where != -1):
            co_name = co_name[:where - 1]
    # 2009 unemployment 16 18 19
    appendlists('2009', 16, 18, 19)
    # 2010 unemployment 20 22 23
    appendlists('2010', 20, 22, 23)
    # 2011 unemployment 24 26 27
    appendlists('2011', 24, 26, 27)
    # 2012 28 30 31
    appendlists('2012', 28, 30, 31)
    # 2013 32 34 35
    appendlists('2013', 32, 34, 35)
    # 2014 36 38 39
    appendlists('2014', 36, 38, 39)
    # 2015 40 42 43
    appendlists('2015', 40, 42, 43)

index = np.arange(len(year))   
df_emp_co = pd.DataFrame(
                       {'Year' : year, 
                        'State' : statename, 
                        'State.Code' : statecode, 
                        'County' : countyname, 
                        'County.Code' : countycode, 
                        'Total.Labor' : totallabor,
                        'Unemployed' : unemployed,
                        'Uemployment.Rate' : unemp_rate },                       
                       index=index)
df_emp_co.to_csv('employment.co.csv', index=False)

index = np.arange(len(year_st))   
df_emp_st = pd.DataFrame(
                       {'Year' : year_st, 
                        'State' : statename_st, 
                        'State.Code' : statecode_st, 
                        'Total.Labor' : totallabor_st,
                        'Unemployed' : unemployed_st,
                        'Uemployment.Rate' : unemp_rate_st },                       
                        index=index)
df_emp_st.to_csv('employment.st.csv', index=False)

    

