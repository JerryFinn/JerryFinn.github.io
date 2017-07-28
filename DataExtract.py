# -*- coding: utf-8 -*-
"""
Created on Tue Jun 13 16:22:01 2017

@author: finnjj
"""
import openpyxl
import pandas as pd
import numpy as np
from pandas import DataFrame, Series

# Columns for SSDI recipients. 
year = []
statename = []
statecode = []
countyname = []
countycode = []
disabledworkers = []

year_st = []
statename_st = []
statecode_st = []
disabledworkers_st = []

#ssdiDict = { 2009 : 'oasdi_sc09.xlsx', 2010 : 'oasdi_sc10.xlsx'}
#ssdiDict = { 2014 : 'oasdi_sc14.xlsx'}
#"""
ssdiDict = { 2009 : 'oasdi_sc09.xlsx', 2010 : 'oasdi_sc10.xlsx', 2011 : 'oasdi_sc11.xlsx',
             2012 : 'oasdi_sc12.xlsx', 2013 : 'oasdi_sc13.xlsx', 2014 : 'oasdi_sc14.xlsx', 
             2015 : 'oasdi_sc15.xlsx' }
#"""
for key_year, value_file in ssdiDict.items():
    wb = openpyxl.load_workbook(value_file)
    #print(wb.get_sheet_names())
    for sheet in wb:
        if ('Table 4 - ' in sheet.title):
            ws = wb.get_sheet_by_name(sheet.title)
            #print(ws.merged_cells)
            state = sheet.title[10:]
            st_code = str(ws.cell(row=5, column=3).value).zfill(2)
            if (not 'Puerto Rico' in state) and (not 'U.S. Virgin Islands' in state):
                #if ('Virginia' == state):
                year_st.append(key_year)
                statename_st.append(state)
                statecode_st.append(st_code)
                disabledworkers_st.append(int(ws.cell(row=5, column=10).value))
                # debug
                #print(ws.rows)
                #print(ws.get_highest_row())
                #print(ws.get_highest_column())
                #print(ws.max_row)
                #print(ws.max_column)
                # debug end
                #for row in ws.iter_rows(): # doesn't work
                #for i in range(6,ws.get_highest_row()):    # deprecated
                for i in range(6,ws.max_row):
                    #print(ws._current_row)     # debug
                    a = ws.cell(row=i, column=1)
                    c = ws.cell(row=i, column=3)
                    j = ws.cell(row=i, column=10)
                    #print(a.value) # debug
                    #print(b.value) # debug
                    if (not a.value) and (not c.value):
                        #print(state) # debug
                        #print(i)     # debug
                        break
                    elif (not j.value):
                        continue
                    elif (not isinstance( j.value, int )):
                        continue
                    else:
                        year.append(key_year)
                        statename.append(state)
                        statecode.append(st_code)
                        countyname.append(ws.cell(row=i, column=1).value)
                        #print(ws.cell(row=i, column=1).value)          # debug
                        countycode.append(str(ws.cell(row=i, column=3).value).zfill(5))
                        disabledworkers.append(int(ws.cell(row=i, column=10).value))
                        #print(ws.cell(row=i, column=10).value)         # debug
                        #print(type(ws.cell(row=i, column=10).value))   # debug
           
index = np.arange(len(year))   
df_ssdi_co = pd.DataFrame(
                       {'Year' : year, 
                        'State' : statename, 
                        'State.Code' : statecode, 
                        'County' : countyname, 
                        'County.Code' : countycode, 
                        'Disabled.Workers' : disabledworkers},
                       index=index)

df_ssdi_co.to_csv('ssdi.co.csv', index=False)

index = np.arange(len(year_st))   
df_ssdi_st = pd.DataFrame(
                       {'Year' : year_st, 
                        'State' : statename_st, 
                        'State.Code' : statecode_st, 
                        'Disabled.Workers' : disabledworkers_st},
                       index=index)

df_ssdi_st.to_csv('ssdi.st.csv', index=False)



           



